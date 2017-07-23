"""
This softtool is used for testing data simple analysis and visualization by using wxPython GUI application, including:
* Using the navigation toolbar
* Adding data to the plot
* Dynamically modifying the plot's properties
* Saving the plot to a file from a menu
Chris Gao (gaole.1989@gmail.com)
Last modified: 06.02.2017
"""
import os
import pprint
import random
import wx
import wx.lib.filebrowsebutton as filebrowse
import xlrd
import openpyxl
from openpyxl.drawing.image import Image

import matplotlib
matplotlib.use('WXAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_wxagg import \
    FigureCanvasWxAgg as FigCanvas, \
    NavigationToolbar2WxAgg as NavigationToolbar
import pylab
import numpy

class BoundControlBox(wx.Panel):
    """ A static box with a couple of radio buttons and a text
        box. Allows to switch between an automatic mode and a 
        manual mode with an associated value.
    """
    def __init__(self, parent, ID, label, initval):
        wx.Panel.__init__(self, parent, ID)
        
        self.value = initval
        
        box = wx.StaticBox(self, -1, label)
        sizer = wx.StaticBoxSizer(box, wx.VERTICAL)
        
        self.radio_auto = wx.RadioButton(self, -1, 
            label="Auto", style=wx.RB_GROUP)
        self.radio_manual = wx.RadioButton(self, -1,
            label="Manual")
        self.manual_text = wx.TextCtrl(self, -1, 
            size=(35,-1),
            value=str(initval),
            style=wx.TE_PROCESS_ENTER)
        
        self.Bind(wx.EVT_UPDATE_UI, self.on_update_manual_text, self.manual_text)
        
        manual_box = wx.BoxSizer(wx.HORIZONTAL)
        manual_box.Add(self.radio_manual, flag=wx.ALIGN_CENTER_VERTICAL)
        manual_box.Add(self.manual_text, flag=wx.ALIGN_CENTER_VERTICAL)
        
        sizer.Add(self.radio_auto, 0, wx.ALL, 10)
        sizer.Add(manual_box, 0, wx.ALL, 10)
        
        self.SetSizer(sizer)
        sizer.Fit(self)
    
    def on_update_manual_text(self, event):
        self.manual_text.Enable(self.radio_manual.GetValue())
    
    def is_auto(self):
        return self.radio_auto.GetValue()
        
    def manual_value(self):
        self.value = self.manual_text.GetValue()
        return self.value


class BarsFrame(wx.Frame):
    """ 
    The main frame of the application
    """
    title = 'Thermal Testing Data Analysis' 

    def __init__(self):
        wx.Frame.__init__(self, None, -1, self.title)
    
        # data needs to be import from the excel
        self.data = []
        self.data_two_scale = []
        self.time = []
        self.plot_legend = ' '
        self.T_limit = []
        self.create_menu() 
        self.create_main_panel() 
        self.draw_figure()

    def create_menu(self):
        self.menubar = wx.MenuBar() 
        
        menu_file = wx.Menu() 
        m_import = menu_file.Append(-1, "&Import\tCtrl-I", "Import file") 
        self.Bind(wx.EVT_MENU, self.on_import, m_import)

        m_expt = menu_file.Append(-1, "&Save plot\tCtrl-S", "Save plot to file") 
        self.Bind(wx.EVT_MENU, self.on_save_plot, m_expt)

        m_export = menu_file.Append(-1,"&Export Excel\tCtrl-E"," Add the pic to Excel")
        self.Bind(wx.EVT_MENU, self.on_export,m_export)

        menu_file.AppendSeparator()
        m_exit = menu_file.Append(-1, "E&xit\tCtrl-X", "Exit")
        self.Bind(wx.EVT_MENU, self.on_exit, m_exit)
        
        menu_help = wx.Menu()
        m_about = menu_help.Append(-1, "&About\tF1", "About the demo")
        self.Bind(wx.EVT_MENU, self.on_about, m_about)
        
        # Import File
        menu_import = wx.Menu()
        m_import = menu_import.Append(-1,"&Import\tCtrl-I","Import the excel")
        self.Bind(wx.EVT_MENU, self.on_import,m_import)
     
        self.menubar.Append(menu_file, "&File")
        self.menubar.Append(menu_help, "&Help")
        self.SetMenuBar(self.menubar)


    def create_main_panel(self):
        """
        Creates the main panel with all the controls on it:
        """

        self.panel = wx.Panel(self)
        
        # Create the mpl Figure and FigCanvas objects. 
        # 10x8 inches, 120 dots-per-inch
        
        self.dpi = 120
        self.fig = Figure((10, 8), dpi=self.dpi)
        self.canvas = FigCanvas(self.panel, -1, self.fig)
        
        self.axes = self.fig.add_subplot(111)
        self.ax2 = self.axes.twinx()
        
        # Bind the 'pick' event for clicking on one of the bars        
        self.canvas.mpl_connect('pick_event', self.on_pick)

        # x y axis control
        self.xmin_control = BoundControlBox(self.panel, -1, "X min", 0)
        self.xmax_control = BoundControlBox(self.panel, -1, "X max", 400)

        # Sheet Selection
        self.sheet_name = wx.StaticText(self.panel,-1,"Sheet") 
        self.sheetlistbox = wx.ListBox(self.panel, -1,(20,40), (90,120), choices = [])
        self.Bind(wx.EVT_LISTBOX,self.SheetEvtListBox,self.sheetlistbox)

        # Column Button /load the column names      
        self.button_column = wx.Button(self.panel,-1,label = u'Column')
        self.Bind(wx.EVT_BUTTON,self.Button_column,self.button_column)

        # Column Selection
        sampleList = []
        self.checklb_column = wx.ListBox(self.panel, -1,(20,40),(140,120), sampleList, wx.LB_SINGLE)
        self.Bind(wx.EVT_LISTBOX, self.EvtListBox, self.checklb_column)

        # Draw Button
        self.drawbutton = wx.Button(self.panel, -1, "Draw")
        self.Bind(wx.EVT_BUTTON, self.on_draw_button, self.drawbutton)

        # Draw Two_scale Button
        self.drawbutton_02 = wx.Button(self.panel, -1, "Draw (Right Y Scale)")
        self.Bind(wx.EVT_BUTTON, self.on_draw_button_02, self.drawbutton_02)        

        # Show Grid
        self.cb_grid = wx.CheckBox(self.panel, -1, 
            "Show Grid",
            style=wx.ALIGN_RIGHT)
        self.Bind(wx.EVT_CHECKBOX, self.on_cb_grid, self.cb_grid)

        # Tprochot Set
        Tprochot_name = wx.Button(self.panel,-1,"Tprochot")
        self.Bind(wx.EVT_BUTTON,self.draw_Tprochot,Tprochot_name)
        self.Tprochot = wx.TextCtrl(self.panel,-1,"0")

        # Clear Button
        clearbutton = wx.Button(self.panel,-1,"Clear")
        self.Bind(wx.EVT_BUTTON,self.clear_button,clearbutton)

        # Create the navigation toolbar, tied to the canvas
        self.toolbar = NavigationToolbar(self.canvas)
        
        # Layout with box sizers        
        self.vbox = wx.BoxSizer(wx.VERTICAL)
        self.vbox.Add(self.canvas, 1, wx.LEFT | wx.TOP | wx.GROW)
        self.vbox.Add(self.toolbar, 0, wx.BOTTOM)
        self.vbox.AddSpacer(5)
        
        self.hbox = wx.BoxSizer(wx.HORIZONTAL)
        flags = wx.ALIGN_LEFT | wx.ALL | wx.ALIGN_CENTER_VERTICAL
     
        self.hbox.AddSpacer(20)
        self.hbox.Add(self.sheet_name,0,border=3,flag=flags)
        self.hbox.AddSpacer(60)
        self.hbox.Add(self.button_column,0,border=3,flag=flags)
        self.hbox.AddSpacer(60)
        self.hbox.Add(self.drawbutton, 0, border=3, flag = flags)
        self.hbox.AddSpacer(10)
        self.hbox.Add(self.drawbutton_02,0,border =3,flag = flags)
        self.hbox.AddSpacer(10)
        self.hbox.Add(Tprochot_name,0,border=3, flag =flags)
        self.hbox.AddSpacer(10)
        self.hbox.Add(self.Tprochot,0,border=3, flag=flags)
        self.hbox.AddSpacer(10)
        self.hbox.Add(clearbutton,0,border=3, flag=flags)               

        self.hbox_02 = wx.BoxSizer(wx.HORIZONTAL)
        self.hbox_02.AddSpacer(10)
        self.hbox_02.Add(self.sheetlistbox,0,border=3, flag=flags)        
        self.hbox_02.AddSpacer(10)
        self.hbox_02.Add(self.checklb_column,0,border=3, flag=flags)
        self.hbox_02.AddSpacer(10)
        self.hbox_02.Add(self.cb_grid,0,border=3, flag=flags)

        self.hbox_02.AddSpacer(10)
        self.hbox_02.Add(self.xmin_control, border=5, flag=wx.ALL)
        self.hbox_02.Add(self.xmax_control, border=5, flag=wx.ALL)        
      
        self.vbox.Add(self.hbox,0, flag = wx.ALIGN_LEFT | wx.TOP)
        self.vbox.AddSpacer(0)
        self.vbox.Add(self.hbox_02,0, flag = wx.ALIGN_LEFT | wx.TOP)       

        self.panel.SetSizer(self.vbox)
        self.vbox.Fit(self)
    
    def create_status_bar(self):
        self.statusbar = self.CreateStatusBar()

    def draw_figure(self):
        """ 
        Redraws the figure
        """
        # clear the axes and redraw the plot anew
           
        self.axes.set_title(u"Data Visualization")
        self.axes.set_xlabel(u"Time(min)")
        self.axes.set_ylabel(u"Temp(Centigrade)")     

        x_lim = self.time
        y_lim = self.data


        if self.plot_legend == ' ':
            pass
        else:
            self.axes.plot(x_lim, y_lim, alpha=0.6, label = self.plot_legend, picker = 5)
            
            if self.xmax_control.is_auto():
                xmax = max(self.time)
            else:
                xmax = int(self.xmax_control.manual_value())
            
            if self.xmin_control.is_auto():            
                xmin = 0
            else:
                xmin = int(self.xmin_control.manual_value())

            self.axes.set_xbound(lower=xmin, upper=xmax)

            legend = self.axes.legend(loc = 'upper right', shadow = True, fontsize = 'small')
            legend.get_frame().set_facecolor('#A4D3EE')

        self.canvas.draw()

    def draw_figure_two_scale(self):
        """ 
        Add the two scale y axis
        """
        x_lim = self.time
        y_lim_02 = self.data        

        if self.plot_legend == ' ':
            pass
        else:     
            self.ax2.plot(x_lim, y_lim_02, 'b',alpha=0.6, label = self.plot_legend, picker = 5)

            legend = self.ax2.legend(loc = 'upper left',shadow = True, fontsize = 'small')
            legend.get_frame().set_facecolor('#A4D3EE')

        self.canvas.draw()

    
    def draw_Tprochot(self,event):
        """ 
        Draw Tprochot on Canvascd 
        """        
        self.T_limit = []
            
        for j in range(0,len(self.time)):
            self.T_limit.append(self.Tprochot.GetValue())

        self.axes.plot(self.time,self.T_limit,'r',alpha=0.6,label = u"Tprochot")
        self.canvas.draw()       

    def clear_button(self,event):
        """ 
        clear canvas
        """  
        self.axes.clear()
        self.data = []

        self.ax2.clear()
        self.axes.set_title("Data Visualization")
        self.axes.set_xlabel(u"Time(min)")
        self.axes.set_ylabel(u"Temp(Centigrade)")

        self.canvas.draw()

    def draw_grid(self):
        self.axes.grid(self.cb_grid.IsChecked())
        self.canvas.draw()

    def EvtListBox(self, event):
        """ 
        Add column value
        """
        self.index = event.GetSelection()

        self.selected_column_value = self.selected_sheetlist.col_values(self.index)
        self.data = self.selected_column_value[1:]
        self.plot_legend = '%s' % self.selected_column_value[0]
            
    def on_cb_grid(self, event):
        self.draw_grid()
    
    def on_draw_button(self, event):
        self.draw_figure()

    def on_draw_button_02(self, event):
        self.draw_figure_two_scale()
    
    def on_pick(self, event):
        """ 
        Display the coords
        """

        line = event.artist

        xdata = line.get_xdata()
        ydata = line.get_ydata()
        ind = event.ind
        plot_points = tuple(zip(xdata[ind],ydata[ind]))

        msg = "Plot coords:\n %s" % plot_points
        
        dlg = wx.MessageDialog(
            self, 
            msg, 
            "Click!",
            wx.OK | wx.ICON_INFORMATION)

        dlg.ShowModal() 
        dlg.Destroy()        

    def on_save_plot(self, event):
        """ 
        Save the plot
        """

        file_choices = "PNG (*.png)|*.png"
        
        dlg = wx.FileDialog(
            self, 
            message="Save plot as...",
            defaultDir=os.getcwd(),
            defaultFile="plot.png",
            wildcard=file_choices,
            style=wx.SAVE)
        
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            self.canvas.print_figure(path, dpi=self.dpi)

    def on_export(self,event):
        """ 
        export the excel
        """

        file_choices = "PNG (*.png)|*.png" 

        export_pic = wx.FileDialog(self,
            message = "Import pic to test report",
            defaultDir = os.getcwd(),
            defaultFile = "plot.png",
            wildcard = file_choices,
            style = wx.OPEN | wx.MULTIPLE | wx.CHANGE_DIR)

        if export_pic.ShowModal() == wx.ID_OK:
            wb=openpyxl.load_workbook(filename = self.file_address)
            paths = export_pic.GetPaths()
            print paths
            new_ws=wb.create_sheet(title='Pics')
            for path in paths:
                img = Image(path)
                new_ws.add_image(img,'A1')
            wb.save('testreport_new.xlsx')
          

    def on_exit(self, event):
        self.Destroy()
        
    def on_about(self, event):
        '''
        click button to see the introduction for this softtool
        '''

        msg = """         
         * Use the matplotlib navigation bar
         * Add Tprochot values to the text box and press Enter (or click "Draw")
         * Show or hide the grid
         * Import Excel, load the sheet and column values.
         * Save the plot to a file using the File menu
         * Click on a plot to receive detailed info
        """
        dlg = wx.MessageDialog(self, msg, "About", wx.OK)
        dlg.ShowModal()
        dlg.Destroy()

    def on_import(self, event):
        '''
        click button to load the excel 
        '''
        dlg = wx.FileDialog(
            self, message="FileBrowse",
            defaultDir=os.getcwd(),
            defaultFile="",
            style=wx.OPEN | wx.MULTIPLE | wx.CHANGE_DIR
            )
        
        if dlg.ShowModal() == wx.ID_OK:
            self.filename = dlg.GetFilename() 
            self.dirname = dlg.GetDirectory()
            f = open(os.path.join(self.dirname, self.filename), 'r')
            self.file_address = os.path.join(self.dirname, self.filename)
            excel_file = xlrd.open_workbook(os.path.join(self.dirname, self.filename))
            excel_file_sheetnames = excel_file.sheet_names()
            for i in excel_file_sheetnames:
                self.sheetlistbox.Append(i)
  
        dlg.Destroy()

    def Button_column(self,event):
        '''
        click button to load the column name from the selected excel sheet
        '''
        self.checklb_column.Clear()
        self.data = []
        self.time = []
        selected_sheet = self.sheetlistbox.GetStringSelection()

        excel_file = xlrd.open_workbook(os.path.join(self.dirname, self.filename))
        self.selected_sheetlist = excel_file.sheet_by_name(selected_sheet)
        selected_column_names = self.selected_sheetlist.row_values(0)
        # x axis Time        
        for i in self.selected_sheetlist.col_values(1)[1:]:
            self.time.append(i)
        # Column list
        for i in selected_column_names:
            self.checklb_column.Append(i)


    def SheetEvtListBox(self,event):
        self.sheet_index = event.GetSelection()


if __name__ == '__main__':
    app = wx.PySimpleApp()
    app.frame = BarsFrame()
    app.frame.Show()
    app.MainLoop()